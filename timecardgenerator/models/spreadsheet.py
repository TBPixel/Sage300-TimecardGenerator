import logging
import openpyxl
import openpyxl.utils as utils

# Application Utilities
from timecardgenerator import tuples


class Spreadsheet( object ):

    def __init__( self, file ):
        assert type( file ) == str, 'file must be a valid string path'

        self.file = file
        self.wb = openpyxl.load_workbook( filename=file, data_only=True )
        self.sheet = self.wb.worksheets[0]

        logging.info( 'Spreadsheet instantiated' )



    def generate( self, employees ):
        """
        Create template spreadsheet
        """
        # Create blank workbook
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet()

        self.sheet = self.wb.worksheets[0]
        # Set worksheet titles
        self.wb.worksheets[0].title = 'Timecard_Header'
        self.wb.worksheets[1].title = 'Timecard_Detail'

        def write_headers( sheet, headers ):
            for key, value in enumerate( headers ):
                index = utils.get_column_letter( key + 1 ) + str( 1 )
                sheet[index] = value

        # Set Worksheet Column Headings
        logging.info( 'Setting Worksheet Column Headings' )
        write_headers( self.wb.worksheets[0], [
            'EMPLOYEE', 'PEREND', 'TIMECARD', 'TCARDDESC', 'TIMESLATE',
            'REUSECARD', 'ACTIVE', 'SEPARATECK', 'PROCESSED', 'CREGHRS',
            'CSHIFTHRS', 'CVACHRSP', 'CVACHRSA', 'CSICKHRSP', 'CSICKHRSP',
            'CCOMPHRSP', 'CCOMPHRSA', 'CVACAMTP', 'CVACAMTA', 'CSICKAMTP',
            'CSICKAMTA', 'CCOMPAMTP', 'CCOMPAMTA', 'CDISIHRSP', 'CDISIHRSA',
            'CDISIAMTP', 'CDISIAMTA', 'LASTNAME', 'FIRSTNAME', 'MIDDLENAME',
            'GREGHRS', 'GSHIFTHRS', 'GVACHRSP', 'GVACHRSA', 'GSICKHRSP',
            'GSICKHRSA', 'GCOMPHRSP', 'GCOMPHRSA', 'GVACAMTP', 'GVACAMTA',
            'GSICKAMTP', 'GSICKAMTA', 'GCOMPAMTP', 'GCOMPAMTA', 'KEYACTION',
            'GDISIHRSP', 'GDISIHRSA', 'GDISIAMTP', 'GDISIAMTA', 'HIREDATE',
            'FIREDATE', 'PARTTIME', 'PAYFREQ', 'OTSCHED', 'COMPTIME',
            'SHIFTSCHED', 'SHIFTNUM', 'WORKPROV', 'STATUS', 'INACTDATE',
            'PROCESSCMD', 'GOTHOURS', 'OTCALCTYPE', 'HRSPERDAY', 'WORKCODE',
            'TOTALJOBS', 'USERSEC', 'WKLYFLSA', 'VALUES', 'OTOVERRIDE',
            'COTHOURS', 'TCDLINES', 'SWJOB', 'SRCEAPPL'
        ])

        write_headers( self.wb.worksheets[1], [
            'EMPLOYEE', 'PEREND', 'TIMECARD', 'LINENUM', 'CATEGORY',
            'EARNDED', 'EARDEDTYPE', 'EARDEDDATE', 'STARTTIME', 'STOPTIME',
            'GLSEG1', 'GLSEG2', 'GLSEG3', 'HOURS', 'CALCMETH', 'LIMITBASE',
            'CNTBASE', 'RATE', 'PAYORACCR', 'EXPACCT', 'LIABACCT', 'OTACCT',
            'SHIFTACCT', 'ASSETACCT', 'OTSCHED', 'SHIFTSCHED', 'SHIFTNUM',
            'WCC', 'TAXWEEKS', 'TAXANNLIZ', 'WEEKLYNTRY', 'ENTRYTYPE',
            'POOLEDTIPS', 'DESC', 'GLSEGID1', 'GLSEGDESC1', 'GLSEGID2',
            'GLSEGDESC2', 'GLSEGID3', 'GLSEGDESC3', 'KEYACTION', 'WORKPROV',
            'PROCESSCMD', 'NKEMPLOYEE', 'NKPEREND', 'NKTIMECARD', 'NKLINENUM',
            'DAYS', 'WCCGROUP', 'VALUES', 'OTHOURS', 'OTRATE', 'SWFLSA',
            'DISTCODE', 'REXPACCT', 'RLIABACCT', 'SWALLOCJOB', 'JOBS',
            'WORKCODE', 'JOBHOURS', 'JOBBASE', 'RCALCMETH', 'RLIMITBASE',
            'RRATEOVER', 'RRATE', 'DEFRRATE'
        ])

        # Fill Timecard_Headers
        key = 2
        for id, employee in employees.items():

            self.write_ws( 0, ( 'A', key ), employee.data.get( 'A' )['value'] ) # ID
            self.write_ws( 0, ( 'B', key ), employee.data.get( 'B' )['value'] ) # Period End
            self.write_ws( 0, ( 'C', key ), employee.data.get( 'C' )['value'] ) # PayPeriod Title

            # Update Key
            key += 1

        self.wb.create_named_range( name='Timecard_Header', worksheet=self.sheet, value='{0}:{1}'.format( self.min_coordinate(), self.max_coordinate() ) )

        # Fill Timecard_Detail
        key = 2
        self.sheet = self.wb.worksheets[1]

        for id, employee in employees.items():

            line = 1

            for shift in employee.hours:

                data       = employee.data.get_all()
                linenumber = line * 1000

                # Dynamically Set constant employee properties
                for column, prop in data.items():
                    self.write_ws( 1, ( column, key ), prop['value'] )

                self.write_ws( 1, ( 'D', key ), linenumber ) # LINENUM
                self.write_ws( 1, ( 'H', key ), shift['date'].date.date() ) # EARNDEDDATE
                self.write_ws( 1, ( 'N', key ), employee.get_shift_sum( shift['hours'] ) ) # HOURS

                # Update Index and line
                key += 1
                line += 1

        self.wb.create_named_range( name='Timecard_Detail', worksheet=self.sheet, value='{0}:{1}'.format( self.min_coordinate(), self.max_coordinate() ) )

        # Save Generated Spreadsheet
        self.wb.save( 'GENERATED-TIMECARDS.xlsx' )
        logging.info( 'DONE!' )



    #        #
    # SHEETS #
    #        #
    def get_cell( self, coordinate ):
        """
        Retrieves a value from a given spreadsheet coordinate

        :param coordinate: str
        :return: any
        """
        return self.sheet[coordinate].value


    def min_coordinate( self ):
        """
        Returns the minimum coordinate found on the active spreadsheet

        :return: str 
        """
        return '{0}{1}'.format( utils.get_column_letter( self.sheet.min_column ), self.sheet.min_row )


    def max_coordinate( self ):
        """
        Returns the maximum coordinate found on the active spreadsheet

        :return: str 
        """
        return '{0}{1}'.format( utils.get_column_letter( self.sheet.max_column ), self.sheet.max_row )


    def loop_sheet( self, loop, callback ):
        """
        loop_sheets( ( start='A1', end='B2' ), lambda coordinate: print( coordinate ) )

        Loops over the given namedtuple coordinates and runs callback on each cell

        :param loop: namedtuple
        :param callback: callable
        """
        assert type( loop ) == tuples.Coordinates, 'Loop coordinates must be of type {0}, instead got {1}'.format( tuples.Coordinates, type( loop ) )
        assert callable( callback ), 'Callback function must be a callable'

        logging.info( 'Looping coordinates {0} : {1}'.format( loop.start, loop.end ) )

        for row in self.sheet[loop.start: loop.end]:

            for cell in row:
                coordinate = utils.coordinate_from_string( cell.coordinate )
                coordinate = tuples.Coordinate( column=coordinate[0], row=coordinate[1] )
                callback( cell, coordinate )


    def write_ws( self, ws, key, value ):
        key = '{0}{1}'.format( key[0], key[1] )
        self.wb.worksheets[ws][key] = value
